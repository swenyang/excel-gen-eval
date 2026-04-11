"""Screenshot generation for Excel sheets.

Uses LibreOffice in headless mode for cross-platform support.
Exports all sheets by converting to PDF (multi-page), then to PNG per page.
"""

from __future__ import annotations

import logging
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from PIL import Image

logger = logging.getLogger(__name__)


def _is_blank_image(img: Image.Image, threshold: float = 0.01) -> bool:
    """Check if an image is near-blank (< threshold fraction of non-white pixels)."""
    import numpy as np
    arr = np.array(img.convert("RGB"))
    non_white = (arr < 250).any(axis=2).sum()
    total = arr.shape[0] * arr.shape[1]
    return (non_white / total) < threshold


class ScreenshotError(RuntimeError):
    """Raised when screenshot generation fails and is required."""


def is_screenshot_available() -> bool:
    """Check if screenshot generation is available (LibreOffice installed)."""
    return _find_libreoffice() is not None


def generate_screenshots(
    excel_path: str | Path,
    required: bool = True,
) -> dict[str, bytes]:
    """Generate PNG screenshots for each sheet in the Excel file.

    Args:
        excel_path: Path to the .xlsx file.
        required: If True, raise ScreenshotError when LibreOffice is missing.

    Returns:
        Dict mapping sheet_name → PNG bytes.
    """
    excel_path = Path(excel_path)
    lo_path = _find_libreoffice()

    if lo_path is None:
        msg = (
            "LibreOffice is required for screenshot generation but was not found. "
            "Install it from https://www.libreoffice.org/download/ "
            "or via: winget install TheDocumentFoundation.LibreOffice"
        )
        if required:
            raise ScreenshotError(msg)
        logger.warning(msg)
        return {}

    # Get sheet names from the workbook
    try:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception:
        sheet_names = []

    try:
        return _generate_via_pdf(excel_path, lo_path, sheet_names)
    except Exception as e:
        logger.warning(f"Screenshot generation failed: {e}")
        return {}


def _find_libreoffice() -> str | None:
    """Find LibreOffice executable on the system."""
    for name in ["libreoffice", "soffice", "soffice.exe"]:
        path = shutil.which(name)
        if path:
            return path

    common_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for p in common_paths:
        if Path(p).exists():
            return p

    return None


def _find_poppler() -> str | None:
    """Find poppler bin directory (needed by pdf2image on Windows)."""
    if shutil.which("pdftoppm"):
        return None  # Already in PATH, no need to specify

    common_paths = [
        r"C:\tools\poppler\poppler-24.08.0\Library\bin",
        r"C:\tools\poppler\Library\bin",
    ]
    # Also scan C:\tools\poppler for any version
    poppler_root = Path(r"C:\tools\poppler")
    if poppler_root.exists():
        for bin_dir in poppler_root.rglob("bin"):
            if (bin_dir / "pdftoppm.exe").exists():
                return str(bin_dir)

    for p in common_paths:
        if Path(p).exists() and (Path(p) / "pdftoppm.exe").exists():
            return p

    return None


def _generate_via_pdf(
    excel_path: Path,
    lo_path: str,
    sheet_names: list[str],
) -> dict[str, bytes]:
    """Convert Excel → PDF (all sheets) → PNG per page."""
    screenshots: dict[str, bytes] = {}

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)

        # Step 0: Create a trimmed copy to avoid bloated PDFs from empty columns/rows
        trimmed_path = _trim_excel_for_screenshot(excel_path, tmpdir_path)

        # Step 1: Convert Excel to PDF (all sheets become pages)
        cmd = [
            lo_path,
            "--headless",
            "--calc",
            "--convert-to", "pdf",
            "--outdir", str(tmpdir_path),
            str(trimmed_path),
        ]

        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=120,
        )

        pdf_files = sorted(tmpdir_path.glob("*.pdf"))
        if not pdf_files:
            logger.warning(f"LibreOffice PDF conversion produced no output: {result.stderr}")
            # Fallback to single PNG export
            return _generate_single_png(excel_path, lo_path, sheet_names)

        pdf_path = pdf_files[0]

        # Step 2: Convert PDF pages to PNGs
        try:
            screenshots = _pdf_to_pngs(pdf_path, sheet_names)
        except Exception as e:
            logger.warning(f"PDF to PNG conversion failed ({e}), falling back to single PNG")
            return _generate_single_png(excel_path, lo_path, sheet_names)

    return screenshots


def _trim_excel_for_screenshot(excel_path: Path, tmpdir: Path) -> Path:
    """Create a trimmed copy of an Excel file, removing trailing empty columns/rows.

    This prevents LibreOffice from generating hundreds of PDF pages for files
    where max_column=16384 due to stray formatting on empty cells.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(str(excel_path))
        needs_trim = False

        for ws in wb.worksheets:
            # Find the last column with actual data
            last_data_col = 0
            last_data_row = 0
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 500),
                                     max_col=min(ws.max_column or 0, 500)):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip():
                        last_data_col = max(last_data_col, cell.column)
                        last_data_row = max(last_data_row, cell.row)

            # Unmerge mega-merged cells that extend far beyond the data area
            # (e.g., A16:XFD16 spanning 16384 columns). These cause LibreOffice
            # to generate hundreds of horizontal PDF pages.
            if last_data_col > 0:
                merges_to_fix = []
                for merge_range in list(ws.merged_cells.ranges):
                    if merge_range.max_col > last_data_col * 2:
                        merges_to_fix.append(merge_range)
                for merge_range in merges_to_fix:
                    needs_trim = True
                    ws.unmerge_cells(str(merge_range))
                    # Re-merge only the data-area portion if it spans multiple rows
                    if merge_range.min_row != merge_range.max_row or merge_range.min_col != merge_range.max_col:
                        new_max_col = min(merge_range.max_col, last_data_col + 1)
                        if new_max_col > merge_range.min_col:
                            ws.merge_cells(
                                start_row=merge_range.min_row, start_column=merge_range.min_col,
                                end_row=merge_range.max_row, end_column=new_max_col,
                            )
                    logger.debug("Fixed mega-merge %s → capped to col %d", merge_range, last_data_col + 1)

            # If Excel reports far more columns than have data, trim is needed
            if ws.max_column and last_data_col > 0 and ws.max_column > last_data_col * 2:
                needs_trim = True
                # Delete columns beyond data range (from right to left)
                if ws.max_column > last_data_col + 1:
                    ws.delete_cols(last_data_col + 1, ws.max_column - last_data_col)

            # Trim trailing empty rows too
            if ws.max_row and last_data_row > 0 and ws.max_row > last_data_row * 2:
                needs_trim = True
                if ws.max_row > last_data_row + 1:
                    ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

        if needs_trim:
            trimmed = tmpdir / f"trimmed_{excel_path.name}"
            wb.save(str(trimmed))
            wb.close()
            logger.info("Trimmed Excel for screenshot: %s", excel_path.name)
            return trimmed

        wb.close()
    except Exception as e:
        logger.debug("Could not trim Excel for screenshot: %s", e)

    return excel_path


def _pdf_to_pngs(pdf_path: Path, sheet_names: list[str]) -> dict[str, bytes]:
    """Convert PDF pages to PNG images.

    Limits conversion to a sensible number of pages. Wide Excel sheets can
    generate hundreds of print pages in a PDF — we take enough to cover
    each sheet's first page but cap total to avoid excessive processing.
    """
    # Heuristic: at most 3 pages per sheet, max 50 total
    max_pages = min(max(len(sheet_names) * 3, 6), 50)

    try:
        from pdf2image import convert_from_path
        poppler_path = _find_poppler()
        kwargs = {"dpi": 150, "last_page": max_pages}
        if poppler_path:
            kwargs["poppler_path"] = poppler_path
        images = convert_from_path(str(pdf_path), **kwargs)
        result: dict[str, bytes] = {}
        for i, img in enumerate(images):
            # Skip near-blank pages (< 1% non-white pixels)
            if _is_blank_image(img):
                continue
            name = sheet_names[i] if i < len(sheet_names) else f"Page_{i + 1}"
            import io
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result[name] = buf.getvalue()
        return result
    except ImportError:
        pass

    # Fallback: use Pillow directly (only works if Pillow has PDF plugin)
    try:
        import io
        result = {}
        img = Image.open(str(pdf_path))
        page = 0
        while True:
            try:
                img.seek(page)
                name = sheet_names[page] if page < len(sheet_names) else f"Sheet_{page + 1}"
                buf = io.BytesIO()
                img.convert("RGB").save(buf, format="PNG")
                result[name] = buf.getvalue()
                page += 1
            except EOFError:
                break
        if result:
            return result
    except Exception:
        pass

    # Last resort: read PDF as single image
    logger.info("Multi-page PDF extraction not available, using single-page fallback")
    return {}


def _generate_single_png(
    excel_path: Path,
    lo_path: str,
    sheet_names: list[str],
) -> dict[str, bytes]:
    """Fallback: export as single PNG (first sheet only)."""
    screenshots: dict[str, bytes] = {}

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        cmd = [
            lo_path, "--headless", "--calc",
            "--convert-to", "png",
            "--outdir", str(tmpdir_path),
            str(excel_path),
        ]
        subprocess.run(cmd, capture_output=True, text=True, timeout=120)

        png_files = sorted(tmpdir_path.glob("*.png"))
        for i, png_file in enumerate(png_files):
            name = sheet_names[i] if i < len(sheet_names) else f"Sheet_{i + 1}"
            screenshots[name] = png_file.read_bytes()

    return screenshots
