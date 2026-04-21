"""Excel file parser — extracts CSV, formulas, charts, and formatting metadata."""

from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart
from openpyxl.utils import get_column_letter

from ..models import (
    ChartInfo,
    FormatInfo,
    FormulaInfo,
    PreparedData,
    SheetData,
)

# Row truncation thresholds
MAX_ROWS_FULL = 500
HEAD_ROWS = 100
TAIL_ROWS = 50


def parse_excel(
    excel_path: str | Path,
    grounding_data: str = "",
    user_prompt: str = "",
) -> PreparedData:
    """Parse an Excel file and extract all structured data for evaluation.

    Returns a PreparedData object with sheets, formulas, charts,
    formatting metadata, and cross-sheet references.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Check for oversized pivot caches or other bloat that causes openpyxl
    # to hang. If the uncompressed xlsx content exceeds a threshold, skip
    # features that require full (non-read_only) workbook parsing.
    heavy = _is_heavy_xlsx(excel_path)
    if heavy:
        import logging
        logging.getLogger(__name__).warning(
            "Large xlsx internals detected (%s), using lightweight parsing", excel_path.name
        )

    if heavy:
        # Lightweight path: skip formatting/cross-refs that need full parse
        wb_values = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
        sheets = _extract_sheets(excel_path, lightweight=True)
        formulas = []  # Can't extract formulas in read_only
        charts = []
        formatting = FormatInfo()
        cross_refs = []
        wb_values.close()
    else:
        wb = openpyxl.load_workbook(str(excel_path), data_only=False)
        wb_values = openpyxl.load_workbook(str(excel_path), data_only=True)
        sheets = _extract_sheets(excel_path)
        formulas = _extract_formulas(wb, wb_values)
        charts = _extract_charts(wb)
        formatting = _extract_formatting(wb)
        cross_refs = _extract_cross_sheet_refs(wb)
        wb.close()
        wb_values.close()

    return PreparedData(
        sheets=sheets,
        formulas=formulas,
        charts=charts,
        formatting=formatting,
        cross_sheet_refs=cross_refs,
        screenshots={},  # Populated separately by screenshot module
        grounding_data=grounding_data,
        user_prompt=user_prompt,
    )


def _is_heavy_xlsx(path: Path, threshold_mb: float = 20.0) -> bool:
    """Check if an xlsx file has oversized internal components.

    Some AI-generated Excel files contain huge pivot caches (50-100MB+)
    that cause openpyxl to hang when loaded in non-read_only mode.
    """
    import zipfile
    try:
        with zipfile.ZipFile(path) as zf:
            for info in zf.infolist():
                if info.file_size > threshold_mb * 1024 * 1024:
                    return True
    except Exception:
        pass
    return False


def _extract_sheets(excel_path: Path, lightweight: bool = False) -> list[SheetData]:
    """Export each sheet to CSV text.

    When *lightweight* is False (default), reads cell number_format via
    openpyxl (full mode) for display formatting.  When True, uses
    openpyxl read_only mode which skips pivot caches and other heavy
    internal data but still preserves number_format for display values.
    """
    xls = pd.ExcelFile(excel_path)
    sheets: list[SheetData] = []

    wb_display = openpyxl.load_workbook(
        str(excel_path), data_only=True, read_only=lightweight,
    )

    for sheet_name in xls.sheet_names:
        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=0)
        row_count, col_count = df_raw.shape

        # Check if sheet is hidden
        is_hidden = False
        if sheet_name in wb_display.sheetnames:
            ws_check = wb_display[sheet_name]
            is_hidden = getattr(ws_check, 'sheet_state', 'visible') in ("hidden", "veryHidden")

        # Build display-formatted DataFrame from openpyxl
        if sheet_name in wb_display.sheetnames:
            ws = wb_display[sheet_name]
            rows_data = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
                if row_idx == 0:
                    headers = [_display_value(cell) for cell in row]
                    # Deduplicate headers (pandas requires unique column names)
                    seen: dict[str, int] = {}
                    for j, h in enumerate(headers):
                        if not h:
                            headers[j] = f"Column_{j+1}"
                        if headers[j] in seen:
                            seen[headers[j]] += 1
                            headers[j] = f"{headers[j]}_{seen[headers[j]]}"
                        else:
                            seen[headers[j]] = 0
                    continue
                rows_data.append([_display_value(cell) for cell in row])

            df = pd.DataFrame(rows_data, columns=headers) if headers else df_raw
        else:
            df = df_raw

        # Drop trailing all-empty rows (openpyxl may include 1M+ rows due to formatting)
        df = df.loc[~(df.isna() | (df.astype(str).str.strip() == "")).all(axis=1)]

        # Drop trailing all-empty columns (openpyxl may include 16384 cols due to formatting)
        if len(df.columns) > 0:
            last_non_empty = 0
            for i, col in enumerate(df.columns):
                col_vals = df[col]
                if col_vals.notna().any() and not (col_vals.astype(str).str.strip() == "").all():
                    last_non_empty = i
            if last_non_empty < len(df.columns) - 1:
                df = df.iloc[:, :last_non_empty + 1]
                col_count = len(df.columns)

        # Use actual DataFrame row count for truncation (openpyxl df may differ from pd.read_excel)
        actual_rows = len(df)
        truncated = False
        if actual_rows > MAX_ROWS_FULL:
            head = df.head(HEAD_ROWS)
            tail = df.tail(TAIL_ROWS)
            truncated_count = actual_rows - HEAD_ROWS - TAIL_ROWS
            marker = pd.DataFrame(
                {col: [f"[... {truncated_count} rows truncated ...]"] for col in df.columns},
                index=[0],
            )
            df = pd.concat([head, marker, tail], ignore_index=True)
            truncated = True

        csv_text = df.to_csv(index=False)
        sheets.append(SheetData(
            name=sheet_name,
            csv_text=csv_text,
            row_count=row_count,
            col_count=col_count,
            truncated=truncated,
            hidden=is_hidden,
        ))

    wb_display.close()
    return sheets


def _display_value(cell) -> str:
    """Format a cell value based on its Excel number_format.

    Produces the value as the user would see it in Excel, not the raw
    underlying value.
    """
    val = cell.value
    if val is None:
        return ""

    fmt = cell.number_format or "General"

    # Date/time formats — detect by common Excel format tokens
    if isinstance(val, datetime):
        if any(tok in fmt.lower() for tok in ["h:", "hh:", "ss", "am/pm"]):
            return val.strftime("%Y-%m-%d %H:%M:%S")
        else:
            return val.strftime("%Y-%m-%d")

    # Numeric — format to match Excel display
    if isinstance(val, (int, float)):
        num = float(val)

        # Percentage format (e.g., "0%", "0.00%")
        if "%" in fmt:
            pct_val = num * 100
            # Count decimal places in percentage format
            if "." in fmt:
                decimal_part = fmt.split(".")[-1].rstrip("%")
                decimals = len(decimal_part.replace("0", "X").replace("#", "X"))
            else:
                decimals = 0
            if decimals == 0:
                return f"{int(round(pct_val))}%"
            return f"{round(pct_val, decimals):.{decimals}f}%"

        use_thousands = "," in fmt or "#,#" in fmt

        # Determine decimal places
        if "." in fmt:
            decimal_part = fmt.split(".")[-1]
            decimals = len(decimal_part.replace("0", "X").replace("#", "X").split(";")[0].rstrip(")%"))
            decimals = min(decimals, 10)
        elif fmt == "General":
            # General format: show value naturally, strip trailing zeros
            # Excel shows 1.0 as "1", 3.14 as "3.14", 5923912.0 as "5923912"
            num = float(val)
            if num == int(num):
                return str(int(num))
            # Strip trailing zeros: 3.140000 → 3.14
            return f"{num:.10g}"
        else:
            decimals = 0

        rounded = round(float(val), decimals)

        # Format with or without thousand separators
        if decimals == 0:
            int_val = int(rounded)
            if use_thousands:
                return f"{int_val:,}"
            return str(int_val)
        else:
            if use_thousands:
                return f"{rounded:,.{decimals}f}"
            return f"{rounded:.{decimals}f}"

    return str(val)


def _extract_formulas(
    wb: openpyxl.Workbook,
    wb_values: openpyxl.Workbook,
) -> list[FormulaInfo]:
    """Extract all formulas with their computed values."""
    formulas: list[FormulaInfo] = []

    for ws in wb.worksheets:
        ws_values = wb_values[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    coord = cell.coordinate
                    computed = ws_values[coord].value
                    has_error = isinstance(computed, str) and computed.startswith("#")

                    formulas.append(FormulaInfo(
                        cell=coord,
                        sheet=ws.title,
                        formula=cell.value,
                        computed_value=computed,
                        has_error=has_error,
                    ))

    return formulas


def _extract_charts(wb: openpyxl.Workbook) -> list[ChartInfo]:
    """Extract chart metadata from all sheets."""
    charts: list[ChartInfo] = []
    chart_type_map = {
        BarChart: "bar",
        LineChart: "line",
        PieChart: "pie",
        AreaChart: "area",
        ScatterChart: "scatter",
    }

    for ws in wb.worksheets:
        for chart in ws._charts:
            chart_type = "unknown"
            for cls, name in chart_type_map.items():
                if isinstance(chart, cls):
                    chart_type = name
                    break

            title = None
            if chart.title:
                title = str(chart.title)

            data_range = None
            if chart.series:
                try:
                    data_range = str(chart.series[0].val.numRef.f)
                except (AttributeError, IndexError):
                    pass

            has_legend = chart.legend is not None
            has_axis_labels = False
            if hasattr(chart, "x_axis") and chart.x_axis.title:
                has_axis_labels = True
            elif hasattr(chart, "y_axis") and chart.y_axis.title:
                has_axis_labels = True

            charts.append(ChartInfo(
                sheet=ws.title,
                chart_type=chart_type,
                title=title,
                data_range=data_range,
                has_legend=has_legend,
                has_axis_labels=has_axis_labels,
            ))

    return charts


def _extract_formatting(wb: openpyxl.Workbook) -> FormatInfo:
    """Extract workbook-level formatting metadata (visible sheets only)."""
    fonts_used: set[str] = set()
    colors_used: set[str] = set()
    merged_ranges: list[str] = []
    conditional_rules: list[str] = []
    frozen_panes: dict[str, str] = {}
    has_cf = False
    has_bold_headers = False
    bordered_cells = 0
    total_cells = 0
    border_styles: set[str] = set()
    theme_font_name: str | None = None

    for ws in wb.worksheets:
        if ws.sheet_state in ("hidden", "veryHidden"):
            continue

        # Check first 3 rows for bold (title, header, or subheader)
        if ws.max_row and ws.max_row > 0:
            for row_num in range(1, min(4, (ws.max_row or 0) + 1)):
                for cell in ws[row_num]:
                    if cell.font and cell.font.bold:
                        has_bold_headers = True
                        break
                if has_bold_headers:
                    break

        # Fonts, colors, and borders
        for row in ws.iter_rows(max_row=min(ws.max_row or 0, 100)):
            for cell in row:
                total_cells += 1
                if cell.font and cell.font.name:
                    if cell.font.scheme:
                        # Track theme font but don't add duplicates
                        if theme_font_name is None:
                            theme_font_name = cell.font.name
                    else:
                        fonts_used.add(cell.font.name)
                if cell.font and cell.font.color and cell.font.color.rgb:
                    color = str(cell.font.color.rgb)
                    if color != "00000000":
                        colors_used.add(color)
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    color = str(cell.fill.start_color.rgb)
                    if color != "00000000":
                        colors_used.add(color)
                # Border detection
                if cell.border:
                    has_any = False
                    for side in ("left", "right", "top", "bottom"):
                        b = getattr(cell.border, side)
                        if b and b.style:
                            has_any = True
                            border_styles.add(b.style)
                    if has_any:
                        bordered_cells += 1

        # Merged cells
        for merged_range in ws.merged_cells.ranges:
            merged_ranges.append(f"{ws.title}!{merged_range}")

        # Conditional formatting
        for cf_rule in ws.conditional_formatting:
            has_cf = True
            conditional_rules.append(f"{ws.title}: {cf_rule}")

        # Frozen panes
        if ws.freeze_panes:
            frozen_panes[ws.title] = str(ws.freeze_panes)

    # Build border summary
    has_borders = bordered_cells > 0
    border_summary = ""
    if total_cells > 0:
        border_pct = bordered_cells / total_cells * 100
        if border_pct > 80:
            border_summary = f"All data cells have borders ({', '.join(sorted(border_styles))} style)"
        elif border_pct > 20:
            border_summary = f"Some cells have borders ({border_pct:.0f}%, {', '.join(sorted(border_styles))} style)"
        elif bordered_cells > 0:
            border_summary = f"Few cells have borders ({bordered_cells} cells)"

    # If no explicit fonts found, use the theme font
    if not fonts_used and theme_font_name:
        fonts_used.add(theme_font_name)

    return FormatInfo(
        fonts_used=sorted(fonts_used),
        color_palette=sorted(colors_used)[:20],  # Cap at 20 colors
        has_conditional_formatting=has_cf,
        conditional_format_rules=conditional_rules[:50],  # Cap at 50 rules
        merged_cell_ranges=merged_ranges,
        frozen_panes=frozen_panes,
        has_bold_headers=has_bold_headers,
        has_borders=has_borders,
        border_summary=border_summary,
    )


def _extract_cross_sheet_refs(wb: openpyxl.Workbook) -> list[str]:
    """Find formulas that reference other sheets.

    Detects:
    - Direct references: Sheet2!A1, 'Sheet Name'!A1
    - Defined names that span sheets
    - INDIRECT() references (flagged as potential cross-sheet)
    """
    cross_refs: list[str] = []
    sheet_names = set(ws.title for ws in wb.worksheets)

    # Check defined names for cross-sheet scope
    try:
        defined = wb.defined_names
        # openpyxl versions differ: try .definedName, then iterate directly
        names_iter = getattr(defined, 'definedName', None) or defined.values()
        for name in names_iter:
            try:
                destinations = list(name.destinations)
                if len(destinations) > 1:
                    sheets_involved = [d[0] for d in destinations]
                    cross_refs.append(
                        f"Defined name '{name.name}' spans sheets: {', '.join(sheets_involved)}"
                    )
                elif destinations and name.attr_text and "!" in name.attr_text:
                    cross_refs.append(
                        f"Defined name '{name.name}' → {name.attr_text}"
                    )
            except Exception:
                pass
    except Exception:
        pass

    # Check formulas for direct sheet references
    seen = set()  # Deduplicate (only track unique source→target pairs)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                formula = cell.value

                for other_sheet in sheet_names:
                    if other_sheet == ws.title:
                        continue
                    # Match: SheetName!Cell, 'Sheet Name'!Cell
                    # Excel escapes single quotes in sheet names as ''
                    escaped_name = other_sheet.replace("'", "''")
                    patterns = [
                        f"{other_sheet}!",
                        f"'{other_sheet}'!",
                        f"'{escaped_name}'!",
                    ]
                    for pattern in patterns:
                        if pattern in formula:
                            pair = (ws.title, other_sheet)
                            if pair not in seen:
                                seen.add(pair)
                                cross_refs.append(
                                    f"{ws.title} → {other_sheet} "
                                    f"(e.g., {cell.coordinate}: {formula[:80]})"
                                )
                            break

                # Flag INDIRECT as potential cross-sheet
                if "INDIRECT(" in formula.upper() and (ws.title, "_INDIRECT") not in seen:
                    seen.add((ws.title, "_INDIRECT"))
                    cross_refs.append(
                        f"{ws.title}!{cell.coordinate} uses INDIRECT() "
                        f"(potential cross-sheet: {formula[:80]})"
                    )

    return cross_refs
